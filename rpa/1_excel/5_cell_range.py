from openpyxl import Workbook
from random import *
wb = Workbook() 
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i, randint(0, 100), randint(0,100)])

col_B = ws["B"] # 영어 column 만 가지고 오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # 영어, 수학 colum 함께 가지고 오기
# print(col_range)
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)



row_title = ws[1]
# for cell in row_title:
#     print(cell.value, end=" ")

# row_range = ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

# 줄이 굉장히 많아서 몇번째 줄까지 있는지 모를때
row_range = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # 각 cell의 정보를 가져옴
        xy = coordinate_from_string(cell.coordinate) # 각 정보를 끊어서 출력해준다. A2 => ('A', 2) B2 = ('B',2)
        # print(xy, end=" ")
        print(xy[0], end="") # A ('A', 2)[0]
        print(xy[1], end=" ") # 2 ('A', 2)[1]
    print()

wb.save("sample1.xlsx")